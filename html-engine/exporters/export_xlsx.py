"""DOM-walking HTML -> XLSX exporter.

Opens a spreadsheet artifact in headless chromium and turns each
<table class="sheet"> into a real, editable worksheet via openpyxl:
header rows are bold + shaded, numeric cells (class="num") are written
as real numbers so Excel can sum them, total rows are bold, and the
column widths fit the content.

Usage:
    python3 export_xlsx.py --url http://host/preview/.../sheet-v1.html \\
                            --out  /path/to/sheet.xlsx

The HTML grid is the source of truth — the page's own structure and
classes (.sheet, .sheet-name, .num, .total) drive the workbook, the
same way the DOCX exporter walks the .doc-page.
"""
from __future__ import annotations
import argparse
import json
import re
import sys

import os
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Walk the rendered page into a list of sheets:
#   { sheets: [ { name, headers:[{text,num}], rows:[{total,cells:[{text,num}]}] } ] }
DOM_WALK_JS = r"""
(() => {
  const pick = document.querySelectorAll('table.sheet');
  const tables = pick.length ? pick : document.querySelectorAll('table');
  const cellNum = (c) => c.classList.contains('num')
      || c.getAttribute('data-type') === 'number'
      || (c.querySelector && c.querySelector('.num'));
  const txt = (c) => (c.textContent || '').replace(/\s+/g, ' ').trim();

  const sheets = [];
  let n = 0;
  for (const t of tables) {
    n++;
    let name = '';
    const wrap = t.closest('.sheet-wrap');
    if (wrap) { const nm = wrap.querySelector('.sheet-name'); if (nm) name = txt(nm); }
    if (!name) { const cap = t.querySelector('caption'); if (cap) name = txt(cap); }
    if (!name) name = 'Sheet ' + n;

    const headRow = t.querySelector('thead tr');
    const headers = headRow
      ? Array.from(headRow.querySelectorAll('th,td')).map(c => ({ text: txt(c), num: !!cellNum(c) }))
      : [];

    // Status pills can't be a shape inside a cell, but their meaning
    // carries: export the pill's text in the pill's color with its tint
    // as the fill. Colored numbers (.pos/.neg) carry their font color too.
    const cellData = (c) => {
      const tinted = c.querySelector('.pill, .pos, .neg, .positive, .negative');
      const pill = c.querySelector('.pill');
      return {
        text: txt(c), num: !!cellNum(c),
        color: tinted ? getComputedStyle(tinted).color : '',
        fill: pill ? getComputedStyle(pill).backgroundColor : '',
      };
    };
    const bodyRows = Array.from(t.querySelectorAll('tbody tr')).map(tr => ({
      total: tr.classList.contains('total') || tr.classList.contains('subtotal'),
      cells: Array.from(tr.querySelectorAll('td,th')).map(cellData),
    })).filter(r => r.cells.length);

    // A table with no <tbody> — treat all rows after the first as data.
    if (!bodyRows.length) {
      const rows = Array.from(t.querySelectorAll('tr'));
      const start = headRow ? 1 : 0;
      for (let i = start; i < rows.length; i++) {
        bodyRows.push({
          total: rows[i].classList.contains('total'),
          cells: Array.from(rows[i].querySelectorAll('td,th')).map(cellData),
        });
      }
    }
    sheets.push({ name, headers, rows: bodyRows });
  }
  return { sheets };
})()
"""

HEADER_FILL = PatternFill("solid", fgColor="F5F5F7")
TOTAL_FILL = PatternFill("solid", fgColor="F5F9FE")
INK = "1D1D1F"
MUTED = "6E6E73"
THIN = Side(style="thin", color="E3E4E8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_number(s: str):
    """Turn a cell's text into a number when it clearly is one, else None.
    Tolerates currency, thousands separators, %, and parentheses-negatives."""
    if s is None:
        return None
    t = s.strip()
    if not t:
        return None
    neg = t.startswith("(") and t.endswith(")")
    cleaned = re.sub(r"[,$€£¥%\s()]", "", t)
    if not re.fullmatch(r"-?\d*\.?\d+", cleaned):
        return None
    try:
        val = float(cleaned)
    except ValueError:
        return None
    if neg:
        val = -val
    pct = t.endswith("%")
    if pct:
        # keep the human number (e.g. 61 for "61%"), Excel users expect
        # the figure they see; a real percent format is a v2 refinement.
        pass
    if val == int(val):
        return int(val)
    return round(val, 6)


def rgb_to_hex(value):
    """A computed CSS color (rgb()/rgba()/#hex) → openpyxl 'RRGGBB', or None
    for empty/transparent (so the default styling is kept)."""
    if not value:
        return None
    v = value.strip()
    nums = re.findall(r"-?\d*\.?\d+", v)
    if v.startswith("rgb") and len(nums) >= 3:
        if len(nums) >= 4 and float(nums[3]) == 0:
            return None
        clamp = lambda x: max(0, min(255, int(round(float(x)))))
        # Explicit FF alpha so Excel never reads it as transparent.
        return "FF{:02X}{:02X}{:02X}".format(clamp(nums[0]), clamp(nums[1]), clamp(nums[2]))
    if v.startswith("#"):
        h = v.lstrip("#")
        if len(h) == 3:
            h = "".join(c * 2 for c in h)
        return ("FF" + h).upper() if len(h) == 6 else None
    return None


def cell_number(text: str):
    """Numeric value for a cell by CONTENT, or None to keep it text.

    Detecting numbers by class alone meant any number the agent forgot to
    tag landed as text — which makes Excel flag "number stored as text".
    So parse by content, but keep true codes/IDs as text:
      - zero-padded values (0007, 01) — leading zeros are meaningful
      - very long digit runs (>15) — account/card numbers Excel mangles
    """
    n = parse_number(text)
    if n is None:
        return None
    core = re.sub(r"[,$€£¥%()\s]", "", (text or "").strip())
    if re.match(r"-?0\d", core):
        return None
    if len(re.sub(r"\D", "", core)) > 15:
        return None
    return n


# Cells that are genuinely "no value" — they don't make a numeric column
# text, and they stay text themselves (Excel doesn't flag non-numbers).
SENTINELS = {"", "-", "–", "—", ".", "n/a", "na", "tbd", "tbc", "—", "null"}


def code_columns(rows: list) -> set:
    """Columns to force entirely to TEXT because they're identifiers, not
    quantities — so a code column never goes half text / half number.

    The signal is a zero-padded value (0007) or an over-long digit run
    (account/card numbers) ANYWHERE in the column. Every other column is
    parsed cell-by-cell: real numbers become numbers (no "stored as text"
    warning) and genuine text (N/A, a footnoted value) stays text — which
    is how a real spreadsheet behaves, and what Excel won't flag.
    """
    codes = set()
    for row in rows:
        for ci, c in enumerate(row.get("cells") or [], 1):
            t = (c.get("text") or "").strip()
            if not t or t.lower() in SENTINELS:
                continue
            core = re.sub(r"[,$€£¥%()\s]", "", t)
            if re.match(r"-?0\d", core) or len(re.sub(r"\D", "", core)) > 15:
                codes.add(ci)
    return codes


def sanitize_title(name: str, used: set) -> str:
    # Excel sheet titles: <=31 chars, no []:*?/\
    t = re.sub(r"[\[\]:\*\?/\\]", " ", name or "Sheet").strip()[:31] or "Sheet"
    base, i = t, 2
    while t.lower() in used:
        suffix = f" {i}"
        t = (base[:31 - len(suffix)] + suffix)
        i += 1
    used.add(t.lower())
    return t


def build(data: dict, out_path: str):
    sheets = data.get("sheets", [])
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    suppress = []   # (1-based sheet index, "A2:A3 C2:C3") — text-number ranges

    sheet_i = 0
    for s in sheets:
        ws = wb.create_sheet(title=sanitize_title(s.get("name"), used))
        sheet_i += 1
        widths = {}
        r = 1
        code_cols = code_columns(s.get("rows") or [])

        headers = s.get("headers") or []
        if headers:
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=r, column=ci, value=h.get("text", ""))
                cell.font = Font(bold=True, color=MUTED, size=10)
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="right" if h.get("num") else "left", vertical="center")
                cell.border = BORDER
                widths[ci] = max(widths.get(ci, 0), len(str(h.get("text", ""))))
            ws.freeze_panes = f"A{r + 1}"
            r += 1

        first_data_row = r
        for row in s.get("rows") or []:
            is_total = bool(row.get("total"))
            for ci, c in enumerate(row.get("cells") or [], 1):
                text = c.get("text", "")
                # Code columns stay text (IDs); every other column parses
                # per cell, so real numbers become numbers and genuine text
                # (N/A, footnotes) stays text.
                num = None if ci in code_cols else cell_number(text)
                value = num if num is not None else text
                cell = ws.cell(row=r, column=ci, value=value)
                # Carry status/number color from the preview (pills, gains/
                # losses); plain cells stay ink.
                fhex = rgb_to_hex(c.get("color")) or INK
                cell.font = Font(bold=is_total or ci == 1, color=fhex, size=11)
                cell.border = BORDER
                cell.alignment = Alignment(
                    horizontal="right" if num is not None else "left", vertical="center")
                bg = rgb_to_hex(c.get("fill"))
                if bg:
                    cell.fill = PatternFill("solid", fgColor=bg)
                elif is_total:
                    cell.fill = TOTAL_FILL
                widths[ci] = max(widths.get(ci, 0), len(str(text)))
            r += 1

        # Code columns are text on purpose (to keep zero-padded IDs), but a
        # numeric-looking code like "1001" would still draw Excel's
        # "number stored as text" triangle. Record the range so we can tell
        # Excel to ignore it (injected into the saved file below).
        last_row = r - 1
        if code_cols and last_row >= first_data_row:
            refs = " ".join(f"{get_column_letter(ci)}{first_data_row}:{get_column_letter(ci)}{last_row}"
                            for ci in sorted(code_cols))
            suppress.append((sheet_i, refs))

        for ci, w in widths.items():
            ws.column_dimensions[get_column_letter(ci)].width = min(60, max(10, w + 3))

    if not wb.sheetnames:
        wb.create_sheet(title="Sheet 1")
    wb.save(out_path)
    _inject_ignored_errors(out_path, suppress)


def _inject_ignored_errors(path, suppress):
    """openpyxl can't write <ignoredErrors>, so patch the saved xlsx zip:
    add it to each sheetN.xml so Excel doesn't flag the intentional
    text-number (code) cells."""
    if not suppress:
        return
    frag_for = {f"xl/worksheets/sheet{i}.xml":
                f'<ignoredErrors><ignoredError sqref="{refs}" numberStoredAsText="1"/></ignoredErrors>'
                for i, refs in suppress}
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            frag = frag_for.get(item.filename)
            if frag:
                xml = data.decode("utf-8")
                # ignoredErrors must precede extLst (and close the sheet).
                if "<extLst>" in xml:
                    xml = xml.replace("<extLst>", frag + "<extLst>", 1)
                else:
                    xml = xml.replace("</worksheet>", frag + "</worksheet>", 1)
                data = xml.encode("utf-8")
            zout.writestr(item, data)
    os.replace(tmp, path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--url", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--timeout", type=int, default=30000)
    args = ap.parse_args()

    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch(args=["--no-sandbox"])
        page = browser.new_page(viewport={"width": 1280, "height": 900})
        page.goto(args.url, wait_until="networkidle", timeout=args.timeout)
        page.wait_for_timeout(300)
        data = page.evaluate(DOM_WALK_JS)
        browser.close()

    if not data or not data.get("sheets"):
        print("no <table class=\"sheet\"> content found", file=sys.stderr)
        sys.exit(2)

    build(data, args.out)
    print(json.dumps({"ok": True, "sheets": len(data["sheets"]), "out": args.out}))


if __name__ == "__main__":
    main()
